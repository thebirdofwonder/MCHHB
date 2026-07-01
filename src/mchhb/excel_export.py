"""接種記録表の Excel 出力。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .table_rows import build_output_rows, disease_merge_spans, vaccine_merge_spans
from .schema import ExtractionResult

COLUMN_WIDTHS = [22, 24, 12, 6, 18, 28, 14, 36]

HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFFF00")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _format_date(date_str: Optional[str]) -> Optional[str]:
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%B %d, %Y")
        except ValueError:
            continue
    return date_str


def export_to_excel(
    result: ExtractionResult,
    output_path: Path,
    *,
    child_name: Optional[str] = None,
) -> Path:
    """抽出結果を接種記録表 Excel として出力する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vaccination Record"

    name = child_name or result.child_name or ""
    output_rows = build_output_rows(result)
    templates = [r.template for r in output_rows]

    ws.merge_cells("A1:H1")
    ws["A1"] = "Vaccination Record"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = "Name"
    ws["B2"] = name
    ws.merge_cells("B2:H2")

    headers = [
        "Disease",
        "Vaccine",
        "Dose",
        "No.",
        "Date of Vaccination",
        "Manufacturer / Product",
        "Lot Number",
        "Medical Institution / Doctor",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    start_row = 4
    for i, out_row in enumerate(output_rows):
        row_idx = start_row + i
        t = out_row.template
        entry = out_row.entry

        ws.cell(row=row_idx, column=1, value=t.disease_en)
        ws.cell(row=row_idx, column=2, value=t.vaccine_ja)
        ws.cell(row=row_idx, column=3, value=t.dose_category)
        ws.cell(row=row_idx, column=4, value=t.dose_number)
        ws.cell(row=row_idx, column=5, value=_format_date(entry.date))
        ws.cell(row=row_idx, column=6, value=entry.manufacturer)
        ws.cell(row=row_idx, column=7, value=entry.lot_number)
        ws.cell(row=row_idx, column=8, value=entry.institution)
        if entry.confidence == "low":
            ws.cell(row=row_idx, column=6).fill = HIGHLIGHT_FILL

        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    if output_rows:
        _apply_merges(ws, start_row, templates)
        end_row = start_row + len(output_rows) - 1
        _auto_row_heights(ws, start_row, end_row)

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A4"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _auto_row_heights(ws, start_row: int, end_row: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        max_lines = 1
        for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not cell.value:
                continue
            text = str(cell.value)
            chars_per_line = max(8, int(width * 1.1))
            lines = max(1, (len(text) + chars_per_line - 1) // chars_per_line)
            max_lines = max(max_lines, lines)
        ws.row_dimensions[row_idx].height = max(18, min(120, max_lines * 15))


def _apply_merges(ws, start_row: int, templates) -> None:
    for row_start, row_end in disease_merge_spans(templates):
        ws.merge_cells(
            start_row=start_row + row_start,
            start_column=1,
            end_row=start_row + row_end,
            end_column=1,
        )
    for row_start, row_end in vaccine_merge_spans(templates):
        ws.merge_cells(
            start_row=start_row + row_start,
            start_column=2,
            end_row=start_row + row_end,
            end_column=2,
        )
